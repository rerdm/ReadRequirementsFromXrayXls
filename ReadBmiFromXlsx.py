import os
import re
import sys
import time

import openpyxl
from datetime import datetime

from openpyxl.styles import PatternFill, Alignment


class ExcelColumnIterator:

    def __init__(self, debug=False):

        """
        This Class will create a Requirement Matrix so to see wich Requirement is
        linked to which testcase.
        In TZBund project the Test-steps containing the BMIs, this programm
        will create a goof readable overview to for teh Requirement-matrix.
        The Header shows all linkt to all testcases (doubles are deleted and the
        header is sorted vorm small BMI to high BMI.

        :param file_path: ( Here you need a Xport from Xray ).
        Mandatory columns: Key, Summary, Manual Test Steps ( containing the BMIS).
        """
        self.column_letter_manual_test_steps = None
        self.test_case_repo_name = None
        self.column_letter_test_case_repo = None
        self.debug = debug

        self.sorted_bmi_list_without_doubles = None
        self.column_letter_key = None
        self.column_letter_summary = None
        self.column_letter_manual_test_step = None

        self.bmi_list = []
        self.global_bmi_list = []

        # Regular expression to Extract BMI from the Manual Test Steps
        self.pattern = r'\bBMI-\d{1,4}\b'
        self.test_case_name = ""
        self.row_number = 0
        self.test_key = ""

        # Template of viel wich contains the BMI-analyzing
        self.date = datetime.now().strftime("%Y-%m-%d_%H-%M(%SSek)")
        self.new_xlsx_name = "Testcases_BundID"

        print("""
        ################################################################################################################
        #                                                                                                              #
        #   BundID Requirements-analysis                                                                               #
        #   This programm will parse a Xray-Xlsx and create a detailed overview of the Requirements-Coverage:          #
        #   - You can see which requirements (BMIs) are linked to wich Test-Cases                                      #
        #   - The header of the new generated Xlsx contains all requirements (sorted from small to large)              #
        #   - If on requirement for a Test-Cases is found the programm will set a X in the dedicated column            #                                                                          
        #   - You can also see the number of requirements linked to wich Test-Case                                     #
        #                                                                                                              #
        #   Created by : Rene Erdmann                                                                                  #                                                                                   
        #                                                                                                              #
        #    PRECONDITION - IMPORTANT:                                                                                 # 
        #     1. You have to create a folder withe the name 'FilteredBmiByTestcase' and place at least one             #
        #        valid Xray-Xlsx (export from Xray).                                                                   #
        #        Mandatory Columns: [Key] , [Summary], [Manual Test Steps], [Test Repository Path]                     #
        #     2. You have to create a folder with the name 'XrayXlsx'.                                                 #
        #        In this folder the new Xlsx with the Requirement-analyses will be saved                               #
        #     3. After this preparation you can continue the programm                                                  #
        #                                                                                                              #         
        ################################################################################################################
        """)

        optional_bmi_file_name_description = input(
            f" - 1. Type in the name of the BundID-Suite wich will be analyzed (Optional-Parameter you can leave it"
            f" blank and press enter)\n"
            f"      EXAMPLE  : YYYY-MM-TT_H-M-S_TestCase_BundID_[YOUR-NAME].xlsx  : > "
            )
        print("")

        if optional_bmi_file_name_description == "":

            self.file_path_to_new_xlsx = f"FilteredBmiByTestcase/{self.date}_" \
                                         f"{self.new_xlsx_name}.xlsx"

        if optional_bmi_file_name_description != "":
            self.file_path_to_new_xlsx = f"FilteredBmiByTestcase/{self.date}_" \
                                         f"{self.new_xlsx_name}" \
                                         f"_{optional_bmi_file_name_description}.xlsx"

        self.new_xlsx_delete_amount_of_unnecessary_row = 9

        # Create the Template of the new xlsx
        workbook = openpyxl.Workbook()
        new_sheet = workbook.active

        new_sheet['A1'] = "Key"
        new_sheet['B1'] = "Summary"
        new_sheet['C1'] = "Test Repository Path"
        new_sheet['D1'] = "BMI-Count"

        # Column configuration
        new_sheet.column_dimensions['A'].width = 10
        new_sheet.column_dimensions['B'].width = 120
        new_sheet.column_dimensions['C'].width = 70

        self.bmi_column_for_new_list = 5
        workbook.save(self.file_path_to_new_xlsx)

        self.xray_xlsx_folder_name = "XrayXlsx/"
        self.file_path = self.user_selection_xray_xlsx_from_xray_xlsx_folder()

        self.write_global_bmis_to_header_list_with_testcases()
        self.create_global_bmi_list_for_bmi_header()

    def user_selection_xray_xlsx_from_xray_xlsx_folder(self):

        xray_xlsx_files = os.listdir(self.xray_xlsx_folder_name )
        print(f" - 2. Select the Xray-Xlsx for Requirements-analyses")
        count_xlsx_files = 0

        xray_xlsx_files_array = []

        for xray_xlsx_file in xray_xlsx_files:
            if xray_xlsx_file.endswith(".xlsx"):

                count_xlsx_files = count_xlsx_files + 1
                print(f"     - [{count_xlsx_files}] - {xray_xlsx_file} - Available XRAy-Xlsx files")
                xray_xlsx_files_array.append(xray_xlsx_file)

        if count_xlsx_files == 1:
            xray_xlsx_files_array.append(xray_xlsx_file)
            print(f"    Automatically selected XRAY-Xlsx (Only one is available) : {xray_xlsx_files_array[count_xlsx_files-1]}")
            user_selected_xray_xlsx_path = xray_xlsx_files_array[count_xlsx_files-1]

        else:
            print(f"      Available XLSX-Files in {self.xray_xlsx_folder_name}-Folder")
            user_selected_xray_xlsx_number = input("     +  Choose number of XLSX : ")

            try:
                user_selected_xray_xlsx_path = xray_xlsx_files_array[int(user_selected_xray_xlsx_number-1)]

            except Exception as e:
                print(f"ERROR : XLSX with number {user_selected_xray_xlsx_number} not available - Error-Code : {e}")
                sys.exit()

        print("")

        return self.xray_xlsx_folder_name+user_selected_xray_xlsx_path

    def create_global_bmi_list_for_bmi_header(self, end_row=None):

        column_count_manual_test_steps = 0
        column_count_key = 0
        column_count_summary = 0
        column_count_test_repo = 0

        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active
            max_row = sheet.max_row

            # Read Values of teh first Row
            first_row_values = [cell.value for cell in sheet[1]]

            for key in first_row_values:
                column_count_key = column_count_key + 1
                if key == "Key":
                    self.column_letter_key = key
                    break

            for key in first_row_values:
                column_count_summary = column_count_summary + 1
                if key == "Summary":
                    self.column_letter_summary = key
                    break

            for key in first_row_values:
                column_count_manual_test_steps = column_count_manual_test_steps + 1
                if key == "Manual Test Steps":
                    self.column_letter_manual_test_step = key
                    break

            for key in first_row_values:
                column_count_test_repo = column_count_test_repo + 1
                if key == "Test Repository Path":
                    self.column_letter_test_case_repo = key
                    break

            if self.column_letter_key and\
                    self.column_letter_summary and\
                    self.column_letter_manual_test_step and\
                    self.column_letter_test_case_repo:
                pass

            else:
                print("")
                print(" - ERROR : Column(s) are missing - check you Xlsx-File header (closing programm)")
                print(f"      Mandatory Fields are : [Key] , [Summary], [Manual Test Steps], [Test Repository Path]")

                if not bool(self.column_letter_key):
                    print(f"     - 'Key' column is missing ")

                if not bool(self.column_letter_summary):
                    print(f"     - 'Summary' column is missing ")

                if not bool(self.column_letter_test_case_repo):
                    print(f"     - 'Test Repository Path' column is missing")

                if not bool(self.column_letter_manual_test_step):
                    print(f"     - 'Manual Test Steps' column is missing")

                sys.exit()

            self.column_letter_manual_test_steps = openpyxl.utils.get_column_letter(column_count_manual_test_steps)

            if end_row is None:
                end_row = max_row
            elif end_row > max_row:
                raise ValueError("End row exceeds maximum row in the sheet")

            for self.row_number in range(2, end_row + 1):

                cell_value = sheet[self.column_letter_manual_test_steps + str(self.row_number)].value
                matches = re.findall(self.pattern, str(cell_value))

                numbers_without_bmi_prefix = [elem.replace("BMI-", "") for elem in matches]

                self.global_bmi_list.append(numbers_without_bmi_prefix)

        except FileNotFoundError:
            print("File not found.")

        except ValueError as e:
            print(e)
        except Exception as e:
            print("An error occurred:", e)

        nested_list = self.global_bmi_list

        # Requirements list unpack
        unpacked_list = [item for sublist in nested_list for item in sublist]

        # Delete double Requirements
        delete_double_bmis_list = list(set(unpacked_list))

        # Sort Requirements
        sorted_bmi_list_without_doubles = sorted(delete_double_bmis_list)

        self.sorted_bmi_list_without_doubles = sorted_bmi_list_without_doubles

        return self.sorted_bmi_list_without_doubles

    def iterate_column(self, end_row=None):

        print(" - 3. Starting Requirement-analyses (this can take some time).... ")
        print("")
        time.sleep(3)

        column_count_manual_test_steps = 0
        column_count_key = 0
        column_count_summary = 0
        column_count_test_repo = 0

        try:
            workbook = openpyxl.load_workbook(self.file_path)
            sheet = workbook.active
            max_row = sheet.max_row

            # Read Values of teh first Row
            first_row_values = [cell.value for cell in sheet[1]]

            for key in first_row_values:
                column_count_key = column_count_key + 1
                if key == "Key":
                    self.column_letter_key = key
                    break

            for key in first_row_values:
                column_count_summary = column_count_summary + 1
                if key == "Summary":
                    self.column_letter_summary = key
                    break

            for key in first_row_values:
                column_count_manual_test_steps = column_count_manual_test_steps + 1
                if key == "Manual Test Steps":
                    self.column_letter_manual_test_step = key
                    break

            for key in first_row_values:
                column_count_test_repo = column_count_test_repo + 1
                if key == "Test Repository Path":
                    self.column_letter_test_case_repo = key
                    break

            if self.column_letter_key and \
                    self.column_letter_summary and \
                    self.column_letter_manual_test_step and \
                    self.column_letter_test_case_repo:
                pass
            else:
                print(" - ERROR : Column are Missing - check you Xlsx file")
                print(f"    Mandatory Fields are : Key , Summary, Manual Test Steps, Test Repository Path")
                print(f"    Your Fields are : "
                      f"    - Key ={self.column_letter_key} , "
                      f"    - Summary={self.column_letter_summary}, "
                      f"    - Test Repository Path={bool(self.column_letter_test_case_repo)} "
                      f"    - Manual Test Steps={bool(self.column_letter_manual_test_step)} ")

                sys.exit()

            self.column_letter_key = openpyxl.utils.get_column_letter(column_count_key)
            self.column_letter_summary = openpyxl.utils.get_column_letter(column_count_summary)
            self.column_letter_test_case_repo = openpyxl.utils.get_column_letter(column_count_test_repo)

            if end_row is None:

                end_row = max_row-self.new_xlsx_delete_amount_of_unnecessary_row

            elif end_row > max_row-self.new_xlsx_delete_amount_of_unnecessary_row:
                raise ValueError("End row exceeds maximum row in the sheet")

            for self.row_number in range(2, end_row + 1):

                cell_value = sheet[self.column_letter_manual_test_steps + str(self.row_number)].value
                matches = re.findall(self.pattern, str(cell_value))

                self.test_case_name = sheet[self.column_letter_summary + str(self.row_number)].value
                self.test_key = sheet[self.column_letter_key + str(self.row_number)].value
                self.test_case_repo_name = sheet[self.column_letter_test_case_repo + str(self.row_number)].value

                numbers_without_bmi_prefix = [elem.replace("BMI-", "") for elem in matches]

                self.bmi_list.extend(numbers_without_bmi_prefix)

                self.find_bmi_in_global_bmi_list_and_set_x(
                    test_key=self.test_key,
                    test_case_name=self.test_case_name,
                    row_number=self.row_number,
                    bmi_list_per_testcase_list=self.bmi_list,
                    test_case_repo_name=self.test_case_repo_name
                )

                time.sleep(0.1)

        except FileNotFoundError:
            print("File not found.")
        except ValueError as e:
            print("Value Error")
            print(e)
        except Exception as e:
            print("An error occurred:", e)

        print(f" - 4. Programm successfully finished - file stored : {self.file_path_to_new_xlsx}")

    def find_bmi_in_global_bmi_list_and_set_x(self, test_key, test_case_name, row_number, bmi_list_per_testcase_list, test_case_repo_name):

        count_bmis_per_testcase = 0

        bmi_global_header_items_to_compare = self.sorted_bmi_list_without_doubles

        bmi_list_per_testcase_list_replaced_dopubles = list(set(bmi_list_per_testcase_list))

        try:

            workbook = openpyxl.load_workbook(self.file_path_to_new_xlsx)
            sheet = workbook.active

            if self.debug:  # Logs will only be shown in debug mode

                print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

                print(f"- ROW Number {row_number} "
                      f"- Test-NAME : {test_case_name} "
                      f"- Test-Case-Repo : {test_case_repo_name} "
                      )

                print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

            sheet['A'+str(row_number)] = test_key
            sheet['B'+str(row_number)] = test_case_name
            sheet['C'+str(row_number)] = test_case_repo_name

            for bmi_per_test_case in bmi_list_per_testcase_list:

                for index, bmi_globals in enumerate(bmi_global_header_items_to_compare):

                    if bmi_per_test_case == bmi_globals:

                        count_bmis_per_testcase = count_bmis_per_testcase + 1
                        column_letter_to_paste_x = openpyxl.utils.get_column_letter(index+self.bmi_column_for_new_list)
                        sheet[column_letter_to_paste_x+str(row_number)] = "X"
                        sheet[column_letter_to_paste_x+str(row_number)].alignment = Alignment(horizontal='center', vertical='center')

                        if self.debug: # Logs will only be shown in debug mode

                            print(f"        Test-Key                                : {test_key}")
                            print(f"        Test is linked to requirement           : BMI-{bmi_per_test_case}")
                            print(f"        Coordinates to fill X                   : ['{column_letter_to_paste_x}{index+self.bmi_column_for_new_list}']")

                        if self.debug:
                            print("-----------------------------------------------------------------------------------")

            if self.debug:

                print(f"        BMI-List Per-testcase                \n"
                      f"        - Count {len(bmi_list_per_testcase_list)} \n"
                      f"        - List : {bmi_list_per_testcase_list}")

                print("-----------------------------------------------------------------------------------")

                print(f"        BMI-List Per-TestCase replace double \n"
                      f"        - Count {len(bmi_list_per_testcase_list_replaced_dopubles)} \n"
                      f"        - List : {bmi_list_per_testcase_list_replaced_dopubles}"
                      )

            sheet['D'+str(row_number)] = len(bmi_list_per_testcase_list_replaced_dopubles) #  count_bmis_per_testcase
            sheet['D'+str(row_number)].alignment = Alignment(horizontal='center', vertical='center')

            if len(bmi_list_per_testcase_list_replaced_dopubles) == 0:

                sheet['D'+str(row_number)].fill = PatternFill(start_color="FF0000",
                                                              end_color="FF0000",
                                                              fill_type="solid")

            workbook.save(self.file_path_to_new_xlsx)

        except Exception as e:
             print("An error occurred while writing to Excel:", e)

        self.bmi_list = []

        time.sleep(0.2)

    def write_global_bmis_to_header_list_with_testcases(self):

        bmi_header = self.create_global_bmi_list_for_bmi_header()

        try:
            workbook = openpyxl.load_workbook(self.file_path_to_new_xlsx)
        except FileNotFoundError:
            raise Exception("Die angegebene Excel-Datei wurde nicht gefunden.")
        except Exception as e:
            raise Exception(f"Fehler beim Öffnen der Excel-Datei: {e}")

        # Wählen Sie das Arbeitsblatt aus (angenommen, das Arbeitsblatt heißt "Sheet1")
        sheet = workbook['Sheet']

        # Schreiben Sie die Daten aus dem Array in die Excel-Datei
        for index, element in enumerate(bmi_header):
            ziel_zelle = sheet.cell(row=1, column=self.bmi_column_for_new_list +index)
            ziel_zelle.value = element

        workbook.save(self.file_path_to_new_xlsx)

if __name__ == '__main__':

    excel_iterator = ExcelColumnIterator(debug=False)

    excel_iterator.iterate_column()




